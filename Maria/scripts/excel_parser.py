import openpyxl
from openpyxl_image_loader import SheetImageLoader
from os import path, mkdir


def parse_excel(source_file: str, destination_dir: 'str'):
    cells_with_errors = []
    if not path.exists(destination_dir):
        mkdir(destination_dir)

    # Открываем Excel-файл
    wb = openpyxl.load_workbook(source_file)
    sheet = wb.active
    loader = SheetImageLoader(sheet)

    row_number = 2
    # Проходимся по строкам таблицы начиная со второй строки (предполагается, что первая строка - заголовки)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Извлекаем артикул, изображение и название товара
        article = row[0]
        if article is None:
            continue

        # Получаем координаты клетки, где хранится изображение
        cell = sheet.cell(row=row_number, column=2).coordinate
        row_number += 1

        try:
            image = loader.get(cell)
        except ValueError:
            cells_with_errors.append(cell)
            print(f"Ошибка при парсинге {source_file}. Ячейка {cell}")
            continue
        product_name = row[2]

        # Формируем имя файла для изображения
        image_filename = f"{article}.png"
        image_path = path.join(destination_dir, image_filename)

        # Сохраняем изображение
        if image:
            try:
                image.save(image_path)
            except Exception as e:
                cells_with_errors.append(cell)
                print(f"Ошибка при сохранении. Текст ошибки: {e}")
        else:
            cells_with_errors.append(cell)
            print(f'Нет изображения для товара "{product_name}". Пропуск.')

    if len(cells_with_errors) > 0:
        result = f"+- успешно, но при работе сценария возникли ошибки с ячейками:\n{cells_with_errors}"
        print(result)
        return result
